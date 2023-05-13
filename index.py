# Python program to create
# a file explorer in Tkinter

# import all components
# from the tkinter library
from tkinter import *
import cv2
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt


# import filedialog module
from tkinter import filedialog

# Function for opening the
# file explorer window
def browseFiles():
    global filename
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select a File",
                                          filetypes=(("Text files", "*.xlsx*"), ("all files", "*.*")))

    # Change label contents
    label_file_explorer.configure(text="File Selected: ", fg="blue")
    return filename

	

def browseFiles1():
    global filename1
    filename1 = filedialog.askopenfilename(initialdir="/",
                                           title="Select a File",
                                           filetypes=(("Text files", "*.jpeg*"), ("all files", "*.*")))

    # Change label contents
    label_file_explorer1.configure(text="File Selected", fg="blue")
    return filename1

	
def abc():
    #abc=browseFiles()
    dataframe = openpyxl.load_workbook(filename)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values

    
    
    dataframe2 = pd.read_excel(filename)


    print(dataframe2)
    #from google.colab.patches import cv2_imshow

    # path
    

    # Reading an image in default mode
    #pqr=browseFiles1()
    image = cv2.imread(filename1)


    # Window name in which image is displayed
    #window_name = 'Image'

    # font
    font = cv2.FONT_HERSHEY_TRIPLEX

    # org
    org = (180,310)

    # fontScale
    fontScale = 2

    # Blue color in BGR
    color = (255, 0, 0)

    # Line thickness of 2 px
    thickness = 4

    # Using cv2.putText() method

    for row in range(1, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            val=col[row].value
            
            image = cv2.putText(image, val, org, font,fontScale, color, thickness, cv2.LINE_AA)
            plt.imshow(image)
            plt.axis("off")
            plt.show()
            image=cv2.imread(filename1)
            
            
    plt.imshow(overlay)
    plt.axis("off")
    plt.show()	
																								
# Create the root window
window = Tk()

# Set window title
window.title('File Explorer')

# Set window size
window.geometry("500x500")

#Set window background color
window.config(background = "white")

# Create a File Explorer label
label_file_explorer = Label(window,
                            text = "File Not Selected",
                            width = 20, height = 1,
                            fg = "red")

label_file_explorer1 = Label(window,
                            text = "File Not Selected",
                            width = 20, height = 1,
                            fg = "red")

	
button_explore = Button(window,
						text = "Select Excel File",
						command = browseFiles)

button_explore1 = Button(window,
						text = "Select photo ",
						command = browseFiles1)

button_submit = Button(window,
					text = "Submit",
					command = abc )

button_exit = Button(window,
					text = "Exit",
					command = exit)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns

label_file_explorer.place(x=300  ,y=50 )
label_file_explorer1.place(x=300  ,y=100 )
button_explore.place(x=80  ,y=50 )

button_explore1.place(x=80  ,y=100 )
button_submit.place(x=100  ,y=150 )
button_exit.place(x=300  ,y=150 )


# Let the window wait for any events
window.mainloop()

